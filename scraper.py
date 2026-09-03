#!/usr/bin/env python3
"""
SZPERACZ MIESZKANIOWY — Autonomiczny agent monitorujący ogłoszenia mieszkań na OLX Lublin.
Scrape'uje kategorie, śledzi ceny/promo/odświeżenia/reaktywacje, generuje JSON i Excel.
"""

import requests
# curl_cffi — impersonacja TLS/JA3 prawdziwego Chrome'a. WAF-y OLX potrafią blokować
# po fingerprincie TLS (charakterystyczny dla pythonowego `requests`), mimo poprawnych
# nagłówków. curl_cffi z impersonate="chrome" podszywa się pod TLS przeglądarki.
# Fallback do `requests`, gdy biblioteka niedostępna (np. lokalnie bez instalacji).
try:
    from curl_cffi import requests as cffi_requests
    _HAS_CURL_CFFI = True
except ImportError:
    cffi_requests = None
    _HAS_CURL_CFFI = False
from bs4 import BeautifulSoup
import json
import os
import re
import time
import random
import logging
import threading
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse, parse_qsl, urlencode, urlunparse
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("szperacz-mieszkaniowy")

PROFILES = {
    "mieszkania_lublin": {
        "url": "https://www.olx.pl/nieruchomosci/mieszkania/wynajem/lublin/",
        "label": "Mieszkania na wynajem — Lublin",
        "is_category": True,
    },
}

# Maksymalna akceptowana cena ofert. Oferty z ceną powyżej tej wartości są
# odrzucane na etapie parsowania (cicho — bez logów i bez wpływu na crosscheck
# poza dolnym progiem). Oferty bez ceny (price=None) przepuszczamy.
MAX_PRICE = 10000

# Maksymalna liczba stron paginacji na profil (zabezpieczenie przed zapętleniem).
MAX_PAGES = 50

# ── Stabilne sortowanie wyników ──────────────────────────────────────────────
# Domyślne sortowanie OLX ("trafność") przetasowuje listę między żądaniami:
# wyróżnione rotują, a kolejność organicznych potrafi się zmienić w trakcie
# jednego sweepu. Przy paginacji offsetowej każde przesunięcie listy w górę
# (zniknięcie oferty u góry) wypycha pierwszą ofertę strony N+1 na stronę N,
# którą już pobraliśmy — i oferta ginie. created_at:desc daje porządek, który
# zmienia się tylko przy realnych zmianach na rynku.
STABLE_SORT_PARAM = ("search[order]", "created_at:desc")

# ── Paginacja równoległa ─────────────────────────────────────────────────────
# Strona 1 leci sekwencyjnie (daje header_count i liczbę stron), reszta równolegle.
# 4 wątki skracają sweep z ~2 min do ~20 s, czyli tyle samo razy zwężają okno,
# w którym lista może się pod nami przesunąć. Wyżej nie wchodzimy — seria
# równoległych żądań to dokładnie ten wzorzec, na który reagują WAF-y.
PAGE_WORKERS = 4
PAGE_JITTER  = (0.2, 0.9)   # losowy odstęp przed żądaniem — rozjeżdża start wątków

# ── Weryfikacja "zaginionych" ofert ──────────────────────────────────────────
# Oferty z bazy, których nie było w sweepie, sprawdzamy bezpośrednio po ich URL-u.
# To zamiana wnioskowania ("nie ma jej na liście → pewnie zniknęła") na pomiar.
VERIFY_MISSING_ENABLED = True
VERIFY_WORKERS         = 4
VERIFY_TIMEOUT_S       = 20
VERIFY_JITTER          = (0.2, 0.9)
# Powyżej tego progu (liczbowo lub jako ułamek bazy — patrz SANITY_MAX_MISSING_RATIO)
# nie weryfikujemy pojedynczo: tyle ofert naraz nie znika, więc zepsuty jest sweep,
# a nie oferty. Przy okazji trzyma budżet żądań w ryzach.
VERIFY_MAX_LISTINGS    = 150
# Bezpiecznik na wypadek, gdyby detekcja martwej strony przestała działać po
# zmianie layoutu OLX: oferta utrzymywana przy życiu wyłącznie weryfikacją
# (niewidziana w żadnym sweepie) i tak trafia do archiwum po tylu dniach.
VERIFY_MAX_ALIVE_DAYS  = 7
# Zapora na WYNIK weryfikacji jako całość. Jeśli niemal wszystkie sprawdzone oferty
# wychodzą martwe, to nie rynek się zawalił — to klasyfikator przestał działać
# (scan #142: 143/143 "martwe", bo frazy o nieaktualnym ogłoszeniu jadą w payloadzie
# SPA na KAŻDEJ stronie). Wtedy odrzucamy całą weryfikację i wracamy do ścieżki 2-scan.
VERIFY_MAX_DEAD_RATIO  = 0.85
VERIFY_MIN_SAMPLE      = 20   # poniżej tylu sprawdzeń odsetek nic nie znaczy
# Budżet czasu na całą weryfikację. Gdy OLX zacznie zwlekać z odpowiedziami, 143 oferty
# × timeout / liczba wątków rozciągnęłyby scan na kilkanaście minut. Po przekroczeniu
# budżetu reszta ofert dostaje 'unknown' — czyli ścieżkę 2-scan, a nie fałszywy werdykt.
VERIFY_TIME_BUDGET_S   = 120

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")
EXCEL_PATH = os.path.join(DATA_DIR, "szperacz_mieszkaniowy.xlsx")
JSON_PATH  = os.path.join(DATA_DIR, "dashboard_data.json")

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/134.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:136.0) Gecko/20100101 Firefox/136.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/135.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.3 Safari/605.1.15",
]

# ─── HTTP Session ────────────────────────────────────────────────────────────

# Profil TLS/JA3, pod który podszywa się curl_cffi. "chrome" mapuje na aktualną
# wersję Chrome. Impersonate ustawia SPÓJNY User-Agent + sec-ch-ua + kolejność
# nagłówków pasujące do TLS — dlatego przy curl_cffi NIE nadpisujemy User-Agent
# losową wartością z USER_AGENTS (Chrome-TLS + obcy UA byłby bardziej podejrzany
# niż samo `requests`).
IMPERSONATE_TARGET = "chrome"

# Lista profili impersonacji do ROTACJI przy blokadzie 403. OLX/CloudFront
# potrafi zablokować konkretny odcisk JA3 — brat zmierzył (ten sam IP), że
# `requests` dostaje 403 na każdy request, a `curl_cffi` z impersonacją 200;
# gdyby padł akurat odcisk "chrome", jeden stały profil zostawia nas bez
# fallbacku. Zaczynamy od IMPERSONATE_TARGET, a przy 403 rotujemy na kolejny.
# Używamy stabilnych aliasów (a nie nazw wersjonowanych jak "chrome131"), bo te
# ostatnie znikają między wydaniami curl_cffi — i tak filtrujemy listę do
# profili faktycznie wspieranych przez zainstalowaną wersję (patrz niżej).
IMPERSONATE_PROFILES = [IMPERSONATE_TARGET, "safari", "firefox", "edge"]

# Ponawianie żądań (curl_cffi nie ma HTTPAdapter/Retry z urllib3 — musimy sami).
HTTP_MAX_RETRIES  = 3   # ponowienia na 429/5xx + błędy transportu (poza pierwszą próbą)
HTTP_BACKOFF_BASE = 2   # sekundy; backoff = BASE * 2**(numer_próby)


def _available_impersonate_profiles():
    """IMPERSONATE_PROFILES ograniczone do nazw wspieranych przez zainstalowaną
    wersję curl_cffi (nazwy profili są wersjonowane i bywają usuwane). Zawsze
    zwraca co najmniej [IMPERSONATE_TARGET]; [] gdy curl_cffi niedostępny."""
    if not _HAS_CURL_CFFI:
        return []
    try:
        import typing
        from curl_cffi.requests.impersonate import BrowserTypeLiteral
        supported = set(typing.get_args(BrowserTypeLiteral))
    except Exception:
        # Starsza wersja bez literału — nie filtrujemy, ufamy aliasom.
        return list(IMPERSONATE_PROFILES)
    avail = [p for p in IMPERSONATE_PROFILES if p in supported]
    return avail or [IMPERSONATE_TARGET]

# Wspólny zestaw wyjątków sieciowych dla obu backendów (curl_cffi + requests).
if _HAS_CURL_CFFI:
    try:
        from curl_cffi.requests.exceptions import RequestException as _CffiError
    except ImportError:  # starsze wersje curl_cffi
        from curl_cffi.requests.errors import RequestsError as _CffiError
    NETWORK_ERRORS = (requests.RequestException, _CffiError)
else:
    NETWORK_ERRORS = (requests.RequestException,)


def get_session(impersonate=None):
    # ── Preferowane: curl_cffi z impersonacją TLS Chrome ──
    if _HAS_CURL_CFFI:
        s = cffi_requests.Session(impersonate=impersonate or IMPERSONATE_TARGET)
        # impersonate dostarcza już User-Agent i nagłówki sec-*; dokładamy tylko
        # język (OLX = rynek PL) — reszty nie ruszamy, by nie rozjechać fingerprintu.
        s.headers.update({
            "Accept-Language": "pl-PL,pl;q=0.9,en-US;q=0.8,en;q=0.7",
        })
        return s

    # ── Fallback: requests (gdy curl_cffi niedostępny) ──
    from requests.adapters import HTTPAdapter
    from urllib3.util.retry import Retry
    s = requests.Session()
    s.headers.update({
        "User-Agent": random.choice(USER_AGENTS),
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "pl-PL,pl;q=0.9,en-US;q=0.8,en;q=0.7",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
    })
    retry = Retry(total=3, backoff_factor=2, status_forcelist=[429, 500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retry)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s

# ─── Helpers ─────────────────────────────────────────────────────────────────

def parse_price(text):
    if not text:
        return None
    cleaned = re.sub(r"[^\d]", "", text.split("zł")[0] if "zł" in text else text)
    try:
        return int(cleaned) if cleaned else None
    except ValueError:
        return None

def parse_date_text(text):
    if not text:
        return None, None
    text = text.strip()
    today = datetime.now().strftime("%Y-%m-%d")
    if "Odświeżono" in text or "odświeżono" in text:
        return None, _extract_date(text)
    if "Dzisiaj" in text or "dzisiaj" in text:
        return today, today
    return _extract_date(text), None

def _extract_date(text):
    months_pl = {
        "stycznia":"01","lutego":"02","marca":"03","kwietnia":"04","maja":"05","czerwca":"06",
        "lipca":"07","sierpnia":"08","września":"09","października":"10","listopada":"11","grudnia":"12",
    }
    today = datetime.now()
    tl = text.lower()
    if "dzisiaj" in tl: return today.strftime("%Y-%m-%d")
    if "wczoraj" in tl: return (today - timedelta(days=1)).strftime("%Y-%m-%d")
    for mpl, mnum in months_pl.items():
        if mpl in tl:
            m = re.search(r"(\d{1,2})\s+" + mpl + r"\s+(\d{4})", tl)
            if m: return f"{m.group(2)}-{mnum}-{m.group(1).zfill(2)}"
            m = re.search(r"(\d{1,2})\s+" + mpl, tl)
            if m: return f"{today.year}-{mnum}-{m.group(1).zfill(2)}"
    return None

def extract_listing_id(url):
    m = re.search(r"ID([a-zA-Z0-9]+)\.html", url)
    if m: return m.group(1)
    return url.rstrip("/").split("/")[-1]

# ─── Promoted Detection ───────────────────────────────────────────────────────

def detect_promoted_status(card):
    """
    Wykrywa typ promocji ogłoszenia OLX.
    Typy (od najsilniejszego sygnału):
      top_listing  🔝  — Wyróżnione na górze listy (featured)
      bump         ⬆️  — Podbite jednorazowo na górę
      highlight    ✨  — Podświetlone tło
      urgent       🔥  — Oznaczone jako pilne
      premium      💎  — Strona/konto premium
    """
    signals = []  # lista krotek (typ_sygnału, pewność, sugerowany_promo_type)

    # ── Sygnał 1: URL ─────────────────────────────────────────────────────────
    for link in card.select('a[href*="/d/oferta/"]'):
        href = link.get('href', '')
        # search_reason=search%7Cpromoted → wyróżnione na górze
        if 'search_reason=search%7Cpromoted' in href or 'reason=promoted' in href:
            signals.append(('url_promoted', 1.0, 'top_listing'))
            break
        # push_up / bump w URL
        if 'push_up' in href.lower() or 'bump' in href.lower():
            signals.append(('url_bump', 0.95, 'bump'))
            break
        if 'promoted' in href.lower() and '/d/oferta/' in href and 'search_reason' not in href:
            signals.append(('url_keyword', 0.80, 'top_listing'))
            break

    # ── Sygnał 2: data-testid ─────────────────────────────────────────────────
    testid_map = {
        'adCard-featured':   ('featured_badge',    1.0,  'top_listing'),
        'listing-ad-badge':  ('listing_ad_badge',  0.95, 'top_listing'),
        'adCard-promoted':   ('testid_promoted',   0.95, 'top_listing'),
        'adCard-top':        ('testid_top',        0.95, 'top_listing'),
        'adCard-bump':       ('testid_bump',       0.95, 'bump'),
        'adCard-pushup':     ('testid_pushup',     0.95, 'bump'),
        'ad-badge':          ('testid_ad_badge',   0.90, 'top_listing'),
        'badge-promoted':    ('testid_badge_promo',0.90, 'top_listing'),
        'adCard-urgent':     ('testid_urgent',     0.95, 'urgent'),
        'adCard-premium':    ('testid_premium',    0.95, 'premium'),
        'adCard-highlight':  ('testid_highlight',  0.95, 'highlight'),
    }
    for testid, (sig_name, conf, promo) in testid_map.items():
        if card.select_one(f'[data-testid="{testid}"]'):
            signals.append((sig_name, conf, promo))

    # ── Sygnał 3: atrybuty data-* na karcie ───────────────────────────────────
    if card.get('data-promoted') or card.get('data-featured') or card.get('data-cy-promoted'):
        signals.append(('data_promoted', 1.0, 'top_listing'))
    if card.get('data-bump') or card.get('data-pushup') or card.get('data-push-up'):
        signals.append(('data_bump', 0.95, 'bump'))
    if card.get('data-urgent') or card.get('data-pilne'):
        signals.append(('data_urgent', 0.95, 'urgent'))
    if card.get('data-premium') or card.get('data-vip'):
        signals.append(('data_premium', 0.95, 'premium'))
    if card.get('data-highlight') or card.get('data-highlighted'):
        signals.append(('data_highlight', 0.90, 'highlight'))

    # ── Sygnał 4: tekst badge ─────────────────────────────────────────────────
    #   Mapowanie tekst → typ promocji (porządek ma znaczenie: bardziej specyficzne pierwsze)
    text_to_promo = [
        # top_listing
        (['Wyróżnione', 'Wyróżnione na górze', 'TOP ogłoszenie', 'TOP', 'Promowane', 'Sponsorowane'], 'top_listing', 0.90),
        # bump
        (['Podbite', 'Podbij', 'Odśwież', 'Push Up', 'Bump'], 'bump', 0.90),
        # urgent
        (['Pilne', 'Pilne!', 'PILNE', 'Срочно'], 'urgent', 0.92),
        # premium
        (['Premium', 'VIP', 'Pakiet Premium', 'Konto Premium'], 'premium', 0.92),
        # highlight
        (['Podświetlone', 'Wyróżnione tło', 'Highlighted'], 'highlight', 0.85),
    ]
    found_text_types = set()
    for el in card.select('span, div, p, strong, em, label, [class*="badge"], [class*="label"], [class*="tag"]'):
        txt = el.get_text(strip=True)
        for texts, promo, conf in text_to_promo:
            if txt in texts and promo not in found_text_types:
                signals.append((f'text_{promo}', conf, promo))
                found_text_types.add(promo)
                break

    # ── Sygnał 5: klasy CSS ───────────────────────────────────────────────────
    element_classes = ' '.join(card.get('class', [])).lower()
    css_map = [
        (['top-ad', 'featured', 'promoted', 'wyroznienie', 'top_ad'], 'top_listing', 0.75),
        (['bump', 'push-up', 'pushup', 'boosted'],                    'bump',        0.75),
        (['urgent', 'pilne', 'asap'],                                  'urgent',      0.75),
        (['premium', 'vip', 'gold'],                                   'premium',     0.75),
        (['highlighted', 'highlight', 'bg-accent', 'tlo'],            'highlight',   0.70),
    ]
    for keywords, promo, conf in css_map:
        if any(kw in element_classes for kw in keywords):
            signals.append((f'css_{promo}', conf, promo))

    # ── Wynik ─────────────────────────────────────────────────────────────────
    if not signals:
        return {'is_promoted': False, 'promotion_type': None, 'confidence': 1.0}

    # Wybierz typ o najwyższej pewności; przy remisie priorytet: top_listing > bump > urgent > premium > highlight
    priority = {'top_listing': 5, 'bump': 4, 'urgent': 3, 'premium': 2, 'highlight': 1}
    best = max(signals, key=lambda s: (s[1], priority.get(s[2], 0)))
    promo_type = best[2]
    max_conf   = best[1]

    return {'is_promoted': True, 'promotion_type': promo_type, 'confidence': max_conf}

# ─── Card Parsing ─────────────────────────────────────────────────────────────

DATE_KEYWORDS = [
    "odświeżono","dzisiaj","wczoraj",
    "stycznia","lutego","marca","kwietnia","maja","czerwca",
    "lipca","sierpnia","września","października","listopada","grudnia",
]

def parse_card(card):
    title = ""
    href  = ""
    for link in card.select('a[href*="/d/oferta/"]'):
        txt = link.get_text(strip=True)
        if txt and len(txt) > 3:
            title = txt
            href  = link.get("href", "")
            break
        elif not href:
            href = link.get("href", "")
    if not href:
        return None
    full_url = href if href.startswith("http") else f"https://www.olx.pl{href}"
    if not title:
        m = re.search(r"/oferta/(.+?)-CID", href)
        if m: title = m.group(1).replace("-"," ").title()
    if not title:
        return None

    price_el = card.select_one('[data-testid="ad-price"]')
    price_text = price_el.get_text(strip=True) if price_el else ""

    date_text = ""
    location_text = ""
    for el in card.find_all(["p","span"]):
        txt = el.get_text(strip=True)
        if not txt or len(txt) > 120:
            continue
        tl = txt.lower()
        if any(kw in tl for kw in DATE_KEYWORDS):
            if " - " in txt:
                parts = txt.split(" - ", 1)
                location_text = parts[0].strip()
                date_text     = parts[1].strip()
            elif not date_text:
                date_text = txt
        elif txt in ["Lublin","Lublin, lubelskie"] and not location_text:
            location_text = txt

    img = card.select_one("img")
    image_url = img.get("src","") if img else ""
    promo = detect_promoted_status(card)

    return {
        "title": title,
        "price_text": price_text,
        "price": parse_price(price_text),
        "date_text": date_text,
        "location": location_text,
        "url": full_url,
        "listing_id": extract_listing_id(full_url),
        "image_url": image_url,
        "is_promoted": promo["is_promoted"],
        "promotion_type": promo["promotion_type"],
    }

def parse_listings_from_soup(soup):
    cards = soup.select('[data-cy="l-card"]')
    if not cards:
        cards = soup.select("div.css-19pezs8")
    if not cards:
        seen = set()
        for link in soup.select('a[href*="/d/oferta/"]'):
            href = link.get("href","")
            if href in seen: continue
            seen.add(href)
            container = link
            for _ in range(6):
                p = container.parent
                if not p: break
                if p.select_one('[data-testid="ad-price"]'):
                    container = p
                    break
                container = p
            if container != link:
                cards.append(container)
    listings = []
    for card in cards:
        parsed = parse_card(card)
        if parsed:
            listings.append(parsed)
    # Filtr cenowy — oferty powyżej MAX_PRICE są odrzucane po cichu.
    # Oferty z price=None (np. "Zapytaj o cenę") są przepuszczane.
    listings = [l for l in listings if l.get("price") is None or l["price"] <= MAX_PRICE]
    return listings

# Liczba w nagłówku bywa formatowana ze spacją tysięczną (zwykłą lub niełamliwą):
# "Znaleźliśmy 1 234 ogłoszeń". Wzorzec na samo \d+ zwracał wtedy None i crosscheck
# przechodził bezrefleksyjnie (scan #142: header=None).
_HEADER_RE = re.compile(r"Znaleźliśmy\s+([\d\s\u00a0\u202f]+?)\s*ogłosze")

def get_total_count_from_header(soup):
    for el in soup.find_all(string=_HEADER_RE):
        m = _HEADER_RE.search(el)
        if m:
            digits = re.sub(r"[^\d]", "", m.group(1))
            if digits:
                return int(digits)
    # Nie znaleziono — pokaż, co OLX faktycznie napisał. Bez tego "header=None"
    # (scan #144) nie mówi, czy zmieniło się brzmienie, czy nagłówka po prostu nie ma,
    # a crosscheck po cichu przechodzi, bo `header is None` traktujemy jako PASS.
    for el in soup.find_all(string=re.compile(r"ogłosze")):
        txt = re.sub(r"\s+", " ", str(el)).strip()
        if txt:
            log.info(f"  [HEADER?] nie sparsowano licznika; kandydat: {txt[:120]!r}")
            break
    return None

def _url_with_params(url, params):
    """Zwraca URL z podmienionymi parametrami query (wartość None = usuń parametr).
    Sklejanie stringami się tu wykłada: URL profilu może już mieć query (sortowanie),
    a parametr `page` bywa w środku, nie na końcu."""
    parts = urlparse(url)
    q = [(k, v) for k, v in parse_qsl(parts.query, keep_blank_values=True)]
    for key, val in params:
        q = [(k, v) for k, v in q if k != key]
        if val is not None:
            q.append((key, val))
    return urlunparse(parts._replace(query=urlencode(q)))


def build_start_url(profile_config):
    """URL startowy profilu z wymuszonym stabilnym sortowaniem (najnowsze pierwsze).
    Profil może się wypisać przez `"stable_sort": False`."""
    url = profile_config["url"]
    if not profile_config.get("stable_sort", True):
        return url
    if STABLE_SORT_PARAM[0] in url or "search%5Border%5D" in url:
        return url
    return _url_with_params(url, [STABLE_SORT_PARAM])


def page_url(base_url, page):
    """URL konkretnej strony wyników (strona 1 = bez parametru `page`)."""
    return _url_with_params(base_url, [("page", str(page) if page > 1 else None)])


def get_last_page_number(soup):
    """Najwyższy numer strony widoczny w paginacji, albo None gdy OLX go nie pokazuje.
    Znany jest z góry → strony 2..N pobieramy równolegle zamiast iść po `pagination-forward`."""
    links = soup.select(
        '[data-testid*="pagination"] a[href*="page="], '
        '[data-cy*="pagination"] a[href*="page="]'
    )
    if not links:
        wrap = soup.select_one('[data-testid="pagination-wrapper"], [data-cy="pagination-wrapper"]')
        if wrap:
            links = wrap.select('a[href*="page="]')
    max_page = None
    for a in links:
        m = re.search(r"page=(\d+)", a.get("href", ""))
        if m:
            n = int(m.group(1))
            if max_page is None or n > max_page:
                max_page = n
    return max_page


# ─── Scraping + Crosscheck ───────────────────────────────────────────────────

def _http_get(session, url, timeout, profiles, profile_idx, max_retries=None):
    """GET z ponawianiem (429/5xx + błędy transportu, backoff wykładniczy) i
    ROTACJĄ profilu impersonacji przy 403 (prawdopodobna blokada odcisku TLS —
    ponawianie tym samym odciskiem nie ma sensu, więc budujemy sesję z kolejnym
    profilem). Statusy 4xx inne niż 403 (np. 404/410) przechodzą bez ponawiania.

    Zwraca (resp, session, profile_idx) — session i profile_idx mogą się zmienić
    po rotacji, więc caller musi ich użyć do kolejnych żądań. Gdy wszystkie
    profile/próby padną, podnosi ostatni wyjątek sieciowy (NETWORK_ERRORS) —
    dzięki temu 403 nie jest już cichy, tylko przerywa scan jak każdy błąd HTTP."""
    if max_retries is None:
        max_retries = HTTP_MAX_RETRIES
    last_exc = None
    resp = None
    attempt = 0
    while attempt <= max_retries:
        try:
            resp = session.get(url, timeout=timeout)
        except NETWORK_ERRORS as e:
            last_exc = e
            if attempt >= max_retries:
                break
            wait = HTTP_BACKOFF_BASE * (2 ** attempt)
            log.warning(f"  HTTP transport error ({e}) — retry za {wait}s (próba {attempt+1}/{max_retries})")
            time.sleep(wait)
            attempt += 1
            continue
        status = resp.status_code
        # 403 → rotacja profilu impersonacji (nie ponawiamy tym samym odciskiem TLS)
        if status == 403 and profiles and profile_idx + 1 < len(profiles):
            new_target = profiles[profile_idx + 1]
            log.warning(f"  HTTP 403 — rotacja profilu impersonacji {profiles[profile_idx]!r} → {new_target!r}")
            session = get_session(impersonate=new_target)
            profile_idx += 1
            attempt = 0          # nowy profil = świeża pula prób
            last_exc = None
            continue
        # 429/5xx → przejściowe, backoff i retry tym samym profilem
        if status == 429 or 500 <= status < 600:
            if attempt < max_retries:
                wait = HTTP_BACKOFF_BASE * (2 ** attempt)
                log.warning(f"  HTTP {status} — retry za {wait}s (próba {attempt+1}/{max_retries})")
                time.sleep(wait)
                attempt += 1
                continue
            break               # wyczerpano próby — niżej podnosimy HTTPError
        return resp, session, profile_idx
    # Wyczerpano próby: podnieś ostatni błąd transportu albo HTTPError z odpowiedzi.
    if last_exc is not None:
        raise last_exc
    resp.raise_for_status()
    return resp, session, profile_idx


# Sesje curl_cffi nie są thread-safe — każdy wątek trzyma własną (wraz z numerem
# profilu impersonacji, bo _http_get potrafi go zrotować po 403).
_thread_local = threading.local()


def _thread_http_get(url, profiles, timeout=30, max_retries=None):
    session = getattr(_thread_local, "session", None)
    idx     = getattr(_thread_local, "profile_idx", 0)
    if session is None:
        session = get_session(impersonate=profiles[idx] if profiles else None)
    resp, session, idx = _http_get(session, url, timeout, profiles, idx, max_retries=max_retries)
    _thread_local.session     = session
    _thread_local.profile_idx = idx
    return resp


def _fetch_page_listings(url, profiles):
    """Pobiera i parsuje jedną stronę wyników (wołane z puli wątków)."""
    time.sleep(random.uniform(*PAGE_JITTER))
    resp = _thread_http_get(url, profiles, timeout=30)
    resp.raise_for_status()
    return parse_listings_from_soup(BeautifulSoup(resp.text, "lxml"))


def scrape_profile(profile_key, profile_config, session):
    """Pobiera wszystkie strony wyników profilu.

    Strona 1 idzie sekwencyjnie — daje `header_count` i numer ostatniej strony.
    Strony 2..N lecą RÓWNOLEGLE (PAGE_WORKERS wątków): sweep skraca się z ~2 min
    do ~20 s, a wraz z nim okno, w którym OLX przesuwa listę pod nami. Gdy
    paginacja nie zdradza liczby stron, wracamy do trybu sekwencyjnego.

    Zwraca też diagnostykę kompletności (`pages_scraped`, `pages_expected`,
    `failed_pages`, `empty_pages`) — sanity check odrzuca na jej podstawie sweep
    ucięty w połowie, zamiast zapisać go jako "mniej ofert na rynku".
    """
    start_url   = build_start_url(profile_config)
    profiles    = _available_impersonate_profiles()
    profile_idx = 0
    collected   = {}     # listing_id -> oferta (dedup: pierwsze wystąpienie wygrywa)
    page_counts = {}     # nr strony -> ile ofert sparsowano
    failed_pages = []

    def absorb(page_no, listings):
        page_counts[page_no] = len(listings)
        added = 0
        for l in listings:
            if l["listing_id"] not in collected:
                collected[l["listing_id"]] = l
                added += 1
        return added

    # ── Strona 1: sekwencyjnie (header + rozpoznanie paginacji) ──
    log.info(f"  [{profile_key}] Page 1: {start_url}")
    resp, session, profile_idx = _http_get(session, start_url, 30, profiles, profile_idx)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")
    header_count = get_total_count_from_header(soup)
    last_page    = get_last_page_number(soup)
    absorb(1, parse_listings_from_soup(soup))
    log.info(f"  [{profile_key}] Page 1: {page_counts[1]} listings | header={header_count} "
             f"| ostatnia strona wg paginacji={last_page}")

    # ── Strony 2..N: równolegle ──
    parallel_last = min(last_page or 1, MAX_PAGES)
    if parallel_last > 1:
        log.info(f"  [{profile_key}] Pobieram strony 2-{parallel_last} równolegle ({PAGE_WORKERS} wątki)")
        with ThreadPoolExecutor(max_workers=PAGE_WORKERS) as ex:
            futures = {n: ex.submit(_fetch_page_listings, page_url(start_url, n), profiles)
                       for n in range(2, parallel_last + 1)}
            for n, fut in futures.items():
                try:
                    listings = fut.result()
                except Exception as e:
                    failed_pages.append(n)
                    log.error(f"  [{profile_key}] Page {n}: BŁĄD ({type(e).__name__}: {e})")
                    continue
                added = absorb(n, listings)
                log.info(f"  [{profile_key}] Page {n}: {len(listings)} listings (+{added} nowych)")

    # ── Ogon: OLX pokazuje w paginacji tylko okno numerów — dobieramy dalej sekwencyjnie ──
    page = parallel_last + 1
    while page <= MAX_PAGES and not failed_pages:
        url = page_url(start_url, page)
        try:
            resp, session, profile_idx = _http_get(session, url, 30, profiles, profile_idx)
            resp.raise_for_status()
        except NETWORK_ERRORS as e:
            failed_pages.append(page)
            log.error(f"  [{profile_key}] Page {page} (ogon): BŁĄD ({type(e).__name__}: {e})")
            break
        listings = parse_listings_from_soup(BeautifulSoup(resp.text, "lxml"))
        if not listings:
            break
        added = absorb(page, listings)
        log.info(f"  [{profile_key}] Page {page} (ogon): {len(listings)} listings (+{added} nowych)")
        if added == 0:
            # Ta sama treść co wcześniej (OLX zwraca ostatnią stronę zamiast 404) — koniec.
            break
        page += 1
        time.sleep(random.uniform(1.0, 2.0))

    # ── Diagnostyka kompletności ──
    # "Dziura" to pusta strona, PO KTÓREJ są jeszcze strony z ofertami. Pusty ogon
    # (paginacja zapowiadała stronę, na której realnie nic już nie ma) jest normalny
    # i nie może wywracać scanu.
    non_empty      = [n for n, c in page_counts.items() if c > 0]
    last_non_empty = max(non_empty) if non_empty else 0
    empty_pages    = sorted(n for n, c in page_counts.items() if c == 0 and n < last_non_empty)
    unique        = list(collected.values())
    return {
        "listings":       unique,
        "count":          len(unique),
        "header_count":   header_count,
        "pages_scraped":  len(page_counts),
        "pages_expected": last_page,
        "failed_pages":   sorted(failed_pages),
        "empty_pages":    empty_pages,
    }

# ── Sanity checks (zapora przed pustymi / fałszywymi scanami) ──────────────
# 4 zabezpieczenia chronią przed sytuacją, gdy OLX zwróci CAPTCHA / pustą stronę
# / błąd po stronie sieci i scan zostanie uznany za udany.
SANITY_MIN_COUNT       = 50      # poniżej traktujemy jako uszkodzenie (kategoria zwykle 500-700)
SANITY_MIN_HEADER      = 10      # header_count < 10 to znak, że strona się nie załadowała
# Po zrównolegleniu paginacji pełny sweep trwa ~20s, więc czas przestał być miarą
# kompletności — tę rolę przejęły liczniki stron (failed_pages / pages_expected).
# Zostaje niski próg, żeby złapać natychmiastowy redirect na CAPTCHA.
SANITY_MIN_DURATION_S  = 5
# Realna dzienna zmiana nigdy nie przekroczyła kilku procent (90 dni historii),
# a próg 40% przepuszczał niepełne scany: 2026-06-15 zapisał 429 zamiast ~585.
SANITY_MAX_DROP_RATIO  = 0.20
SANITY_MAX_MISSING_RATIO = 0.25  # >25% bazy znika w jednym scanie (carried_missing) = niepełny scrape
COOLDOWN_AFTER_ANOMALY = 90      # sekundy pauzy przed retry po wykryciu anomalii

def _previous_good_count(profile_key):
    """Czyta ostatni udany count z data/dashboard_data.json (daily_counts)."""
    try:
        with open(JSON_PATH, encoding="utf-8") as f:
            d = json.load(f)
        dc = d.get("profiles", {}).get(profile_key, {}).get("daily_counts", [])
        # Bierzemy ostatni count > 0 (ignorujemy potencjalnie uszkodzone)
        for entry in reversed(dc):
            c = entry.get("count", 0)
            if c and c > 0:
                return c
    except Exception:
        pass
    return None

def _check_sanity(profile_key, result, duration_s, previous_count):
    """Weryfikuje wynik scanu względem 4 warunków zdrowia.
    Zwraca (ok: bool, reasons: list[str])."""
    count  = result.get("count", 0)
    header = result.get("header_count")
    reasons = []
    # 1) count == 0 → ZAWSZE error (niezależnie od header_count)
    if count == 0:
        reasons.append(f"count=0 (puste wyniki)")
    # 2) count < SANITY_MIN_COUNT
    elif count < SANITY_MIN_COUNT:
        reasons.append(f"count={count} < {SANITY_MIN_COUNT} (próg minimum)")
    # 3) header_count < SANITY_MIN_HEADER (gdy znany)
    if header is not None and header < SANITY_MIN_HEADER:
        reasons.append(f"header_count={header} < {SANITY_MIN_HEADER}")
    # 4) czas trwania
    if duration_s is not None and duration_s < SANITY_MIN_DURATION_S:
        reasons.append(f"duration={duration_s:.1f}s < {SANITY_MIN_DURATION_S}s (zbyt szybko)")
    # 5) spadek > 40% vs poprzedni udany scan
    if previous_count and count > 0:
        drop_ratio = (previous_count - count) / previous_count
        if drop_ratio > SANITY_MAX_DROP_RATIO:
            reasons.append(f"spadek {drop_ratio*100:.1f}% vs poprzedni count={previous_count}")
    # 6) sweep ucięty w połowie — strony, których nie udało się pobrać.
    #    Wcześniej taki scan przechodził jako pełnoprawny (pętla po prostu robiła
    #    `break`), a brakujące oferty szły w missing → po dwóch dniach w archiwum.
    failed = result.get("failed_pages") or []
    if failed:
        reasons.append(f"nie pobrano stron: {failed}")
    # 7) mniej stron niż zapowiadała paginacja
    expected = result.get("pages_expected")
    scraped_pages = result.get("pages_scraped")
    if expected and scraped_pages is not None:
        expected_capped = min(expected, MAX_PAGES)
        if scraped_pages < expected_capped:
            reasons.append(f"pobrano {scraped_pages} z {expected_capped} stron paginacji")
    # 8) dziura w środku paginacji (strona bez ofert, choć dalsze mają)
    empty = result.get("empty_pages") or []
    if empty:
        reasons.append(f"puste strony w środku paginacji: {empty}")
    return (len(reasons) == 0, reasons)

def scrape_with_crosscheck(profile_key, profile_config):
    backend = f"curl_cffi (impersonate={IMPERSONATE_TARGET})" if _HAS_CURL_CFFI else "requests (fallback)"
    log.info(f"[SCAN] Crosscheck: {profile_key} | HTTP backend: {backend}")
    prev_count = _previous_good_count(profile_key)
    log.info(f"  [{profile_key}] Poprzedni udany count: {prev_count}")

    # ── Pierwsza próba ──
    t0 = time.time()
    r1 = scrape_profile(profile_key, profile_config, get_session())
    d1 = time.time() - t0
    r1["duration_s"] = round(d1, 1)
    scraped, header = r1["count"], r1["header_count"]
    log.info(f"  [{profile_key}] Próba 1: count={scraped}, header={header}, duration={d1:.1f}s, "
             f"stron={r1.get('pages_scraped')}/{r1.get('pages_expected')}")

    # Sanity check próby 1
    ok1, reasons1 = _check_sanity(profile_key, r1, d1, prev_count)
    if not ok1:
        log.warning(f"[SANITY-FAIL] {profile_key} (próba 1): {' | '.join(reasons1)}")
        log.info(f"[COOLDOWN] {profile_key}: pauza {COOLDOWN_AFTER_ANOMALY}s przed retry...")
        time.sleep(COOLDOWN_AFTER_ANOMALY)
        # ── Retry po cooldown ──
        t0 = time.time()
        r2 = scrape_profile(profile_key, profile_config, get_session())
        d2 = time.time() - t0
        r2["duration_s"] = round(d2, 1)
        log.info(f"  [{profile_key}] Próba 2 (po cooldown): count={r2['count']}, header={r2['header_count']}, duration={d2:.1f}s")
        ok2, reasons2 = _check_sanity(profile_key, r2, d2, prev_count)
        if not ok2:
            # Bezpieczny exit: oznaczamy jako anomaly_detected — main.py NIE zmodyfikuje dashboard_data.json
            log.error(f"[ANOMALY] {profile_key}: scan odrzucony (próba 2 też zła: {' | '.join(reasons2)})")
            r2["crosscheck"] = "anomaly_detected"
            r2["anomaly_reasons"] = reasons2
            r2["previous_good_count"] = prev_count
            return r2
        # Retry udany — kontynuujemy crosscheck na r2
        log.info(f"[RECOVERY] {profile_key}: retry po cooldown przeszedł sanity, kontynuuję crosscheck")
        r1 = r2
        scraped, header = r1["count"], r1["header_count"]

    # ── Crosscheck (header vs scraped) ──
    # OLX miesza ~38% kart Otodom w wynikach kategorii — tolerancja musi to uwzględniać.
    # Dla kategorii: do 50% różnicy to normalne zachowanie OLX.
    tolerance = int(header * 0.50) if (profile_config.get("is_category") and header) else 10
    if header is None or abs(scraped - header) <= tolerance:
        log.info(f"[CROSSCHECK] {profile_key}: PASS (scraped={scraped}, header={header})")
        r1["crosscheck"] = "passed"
        return r1

    log.info(f"[CROSSCHECK] {profile_key}: MISMATCH scraped={scraped} vs header={header}, retrying...")
    time.sleep(random.uniform(3, 5))
    t0 = time.time()
    r2 = scrape_profile(profile_key, profile_config, get_session())
    r2["duration_s"] = round(time.time() - t0, 1)
    c1, c2 = r1["count"], r2["count"]

    # Sanity check na r2 też — żeby best_of_two nie wpychało zerowego scanu
    ok_r2, _ = _check_sanity(profile_key, r2, r2["duration_s"], prev_count)
    if not ok_r2:
        log.warning(f"[CROSSCHECK] {profile_key}: r2 nie przeszło sanity, używam r1")
        r1["crosscheck"] = "best_of_two"
        return r1

    if header is not None:
        if abs(c2-header) < abs(c1-header):
            r2["crosscheck"] = "passed_retry"; return r2
        if c1 == c2:
            r1["crosscheck"] = "consistent"; return r1
    else:
        if c2 > c1:
            r2["crosscheck"] = "no_header_retry"; return r2
    r1["crosscheck"] = "best_of_two"; return r1

# ─── Weryfikacja "zaginionych" ofert ─────────────────────────────────────────
# Oferta może zniknąć z listingu, choć nadal żyje: sweep łapie ją nierówno przy
# paginacji offsetowej, a OLX potrafi zwrócić niepełne wyniki. Zamiast wnioskować
# z nieobecności, pytamy OLX wprost o tę ofertę — jej własnym URL-em.

# Frazy ze stron wygaszonych/usuniętych ogłoszeń. Szukamy ich WYŁĄCZNIE w tekście
# widocznym — OLX to SPA i w <script> jedzie payload z kompletem stringów tłumaczeń,
# więc dopasowanie w surowym HTML wychodziło na każdej stronie (scan #142: 143/143
# ofert uznanych za martwe). Trzymamy pełne zwroty, nie pojedyncze słowa — "zakończone"
# czy "wygasło" trafiają się też w treści żywych ogłoszeń.
DEAD_PAGE_MARKERS = (
    "to ogłoszenie jest już nieaktualne",
    "ogłoszenie jest już nieaktualne",
    "ogłoszenie zostało usunięte",
    "ogłoszenie nie jest już dostępne",
    "ogłoszenie wygasło",
    "nie znaleźliśmy tej strony",
    "strona, której szukasz, nie istnieje",
    "this ad is no longer available",
)

# Elementy, które renderuje żywa strona oferty. Sprawdzamy je SELEKTOREM na drzewie
# DOM (po wycięciu <script>), a nie szukaniem stringa w HTML — z tego samego powodu.
ALIVE_PAGE_SELECTOR = ", ".join((
    '[data-testid="ad-price-container"]',
    '[data-testid="main-ad-price"]',
    '[data-testid="ad_description"]',
    '[data-cy="ad_title"]',
    '[data-cy="ad_description"]',
    '[data-testid="ad-footer-bar-section"]',
))


def _visible_soup(html):
    """DOM bez <script>/<style>/<noscript>/<template> — czyli to, co widzi człowiek."""
    soup = BeautifulSoup(html or "", "lxml")
    for tag in soup(["script", "style", "noscript", "template"]):
        tag.decompose()
    return soup


def classify_offer_page(status_code, final_url, html):
    """Czy strona pojedynczej oferty świadczy o tym, że ogłoszenie żyje?

    Zwraca `(status, powód)`, gdzie status to 'dead' | 'alive' | 'unknown'.
    Powód idzie do logów — bez niego diagnoza błędnej klasyfikacji wymaga zgadywania.

    Do 'dead' schodzimy WYŁĄCZNIE na jednoznaczny sygnał: pomyłka w tę stronę
    archiwizuje żywą ofertę, a przy blokadzie WAF (403 na wszystkim) zarchiwizowałaby
    całą bazę w jednym scanie. Sprzeczne sygnały (żywy layout + fraza o nieaktualnym
    ogłoszeniu) to też 'unknown' — wtedy decyduje stara ścieżka 2-scan.
    """
    if status_code in (404, 410):
        return "dead", f"HTTP {status_code}"
    if status_code != 200:
        return "unknown", f"HTTP {status_code}"
    if final_url and "/oferta/" not in final_url:
        # OLX przekierowuje wygaszone oferty na kategorię lub stronę główną.
        return "dead", f"redirect → {final_url[:80]}"

    soup  = _visible_soup(html)
    text  = re.sub(r"\s+", " ", soup.get_text(" ")).strip().lower()
    dead  = next((m for m in DEAD_PAGE_MARKERS if m in text), None)
    alive = soup.select_one(ALIVE_PAGE_SELECTOR) is not None

    if dead and alive:
        return "unknown", f"sprzeczne sygnały: layout żywy + fraza {dead!r}"
    if dead:
        return "dead", f"fraza: {dead!r}"
    if alive:
        return "alive", "layout żywej oferty"
    return "unknown", f"nierozpoznany layout | {_page_title(html)}"


def _page_title(html):
    """<title> strony — tylko do logów przy statusie 'unknown'."""
    m = re.search(r"<title[^>]*>(.*?)</title>", html or "", re.I | re.S)
    return re.sub(r"\s+", " ", m.group(1)).strip()[:120] if m else ""


def _verify_one_listing(url, profiles):
    """(status, powód) dla jednej oferty. Błąd sieci → 'unknown'."""
    time.sleep(random.uniform(*VERIFY_JITTER))
    try:
        resp = _thread_http_get(url, profiles, timeout=VERIFY_TIMEOUT_S, max_retries=1)
    except NETWORK_ERRORS as e:
        return "unknown", f"{type(e).__name__}: {e}"
    final_url = str(getattr(resp, "url", "") or url)
    return classify_offer_page(resp.status_code, final_url, resp.text or "")


def verify_listings_alive(profile_key, missing):
    """Sprawdza równolegle URL-e ofert nieobecnych w sweepie.
    Zwraca (mapa id→status, statystyki)."""
    profiles = _available_impersonate_profiles()
    statuses = {}
    samples  = {}

    deadline = time.time() + VERIFY_TIME_BUDGET_S

    def _job(listing):
        if time.time() > deadline:
            return listing["id"], ("unknown", "przekroczony budżet czasu weryfikacji")
        try:
            return listing["id"], _verify_one_listing(listing["url"], profiles)
        except Exception as e:
            # Weryfikacja jest dodatkiem — jej wywrotka nie może przewrócić scanu.
            return listing["id"], ("unknown", f"{type(e).__name__}: {e}")

    log.info(f"  [{profile_key}] Weryfikacja {len(missing)} zaginionych ofert po URL "
             f"({VERIFY_WORKERS} wątki)...")
    t0 = time.time()
    with ThreadPoolExecutor(max_workers=VERIFY_WORKERS) as ex:
        for lid, (status, reason) in ex.map(_job, missing):
            statuses[lid] = status
            # Próbki KAŻDEGO werdyktu (nie tylko nierozstrzygniętych) — bez nich
            # diagnoza błędnej klasyfikacji wymaga zgadywania po samych licznikach.
            bucket = samples.setdefault(status, [])
            if len(bucket) < 3:
                bucket.append(f"{lid}: {reason}")

    stats = {
        "checked": len(statuses),
        "alive":   sum(1 for s in statuses.values() if s == "alive"),
        "dead":    sum(1 for s in statuses.values() if s == "dead"),
        "unknown": sum(1 for s in statuses.values() if s == "unknown"),
        "duration_s": round(time.time() - t0, 1),
        "skipped": None,
    }
    log.info(f"  [{profile_key}] Weryfikacja: żywe={stats['alive']} martwe={stats['dead']} "
             f"nierozstrzygnięte={stats['unknown']} ({stats['duration_s']}s)")
    for status in ("dead", "alive", "unknown"):
        for s in samples.get(status, []):
            log.info(f"    [VERIFY-{status.upper()}] {s}")
    return statuses, stats


def verify_missing_for_profile(profile_key, result):
    """Uzupełnia `result` o mapę id→status dla ofert z bazy nieobecnych w sweepie.

    Nie weryfikujemy, gdy zaginionych jest podejrzanie dużo — wtedy zepsuty jest
    sweep, nie oferty (i nie ma sensu wysyłać setek żądań, żeby to potwierdzić).
    """
    if not VERIFY_MISSING_ENABLED:
        return
    if result.get("crosscheck") in ("error", "anomaly_detected") or not result.get("count"):
        return
    data    = load_existing_json()
    current = data.get("profiles", {}).get(profile_key, {}).get("current_listings", [])
    if not current:
        return
    scraped = {l["listing_id"] for l in result["listings"]}
    missing = [l for l in current if l.get("id") not in scraped and l.get("url")]
    if not missing:
        result["verification"] = {}
        result["verification_stats"] = {"checked": 0, "alive": 0, "dead": 0,
                                        "unknown": 0, "skipped": None}
        return

    ratio = len(missing) / len(current)
    if ratio > SANITY_MAX_MISSING_RATIO or len(missing) > VERIFY_MAX_LISTINGS:
        reason = (f"{len(missing)} zaginionych ({ratio*100:.1f}% bazy) — "
                  f"to sygnał niepełnego sweepu, nie odpływu ofert; pomijam weryfikację")
        log.warning(f"  [{profile_key}] {reason}")
        result["verification"] = {}
        result["verification_stats"] = {"checked": 0, "alive": 0, "dead": 0,
                                        "unknown": 0, "skipped": reason}
        return

    statuses, stats = verify_listings_alive(profile_key, missing)

    # ── Zapora na wynik zbiorczy ──
    # Prawie same werdykty "martwa" oznaczają awarię klasyfikatora, nie zawał rynku
    # (scan #142: 143/143 martwych, bo frazy o nieaktualnym ogłoszeniu jadą w payloadzie
    # SPA na każdej stronie). Odrzucamy wtedy CAŁĄ weryfikację — archiwizacją zajmie się
    # stara ścieżka 2-scan, która na taki błąd jest odporna.
    checked = stats["checked"]
    if checked >= VERIFY_MIN_SAMPLE and stats["dead"] / checked > VERIFY_MAX_DEAD_RATIO:
        reason = (f"{stats['dead']}/{checked} sprawdzonych ofert wyszło martwych "
                  f"(> {VERIFY_MAX_DEAD_RATIO*100:.0f}%) — to awaria klasyfikacji, "
                  f"nie odpływ ofert; odrzucam całą weryfikację")
        log.error(f"  [{profile_key}] [VERIFY-ABORT] {reason}")
        stats["skipped"] = reason
        stats["aborted_dead_ratio"] = round(stats["dead"] / checked, 3)
        result["verification"] = {}
        result["verification_stats"] = stats
        return

    result["verification"] = statuses
    result["verification_stats"] = stats


# ─── Excel ───────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1A3A6B")
HEADER_FONT = Font(bold=True, color="FFD700", name="Arial", size=10)
DATA_FONT   = Font(name="Arial", size=10)
UP_FONT     = Font(name="Arial", size=10, color="00B050")
DOWN_FONT   = Font(name="Arial", size=10, color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin",color="D9D9D9"), right=Side(style="thin",color="D9D9D9"),
    top=Side(style="thin",color="D9D9D9"),  bottom=Side(style="thin",color="D9D9D9"),
)

def style_header_row(ws, row, num_cols):
    for col in range(1, num_cols+1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL; c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = THIN_BORDER

def style_data_cell(cell, font=None):
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = Alignment(vertical="center", wrap_text=True)

def load_or_create_workbook():
    if os.path.exists(EXCEL_PATH):
        try: return load_workbook(EXCEL_PATH)
        except Exception as e: log.warning(f"Cannot load Excel: {e}. Creating new.")
    wb = Workbook(); wb.remove(wb.active); return wb

def get_or_create_sheet(wb, name, headers):
    if name in wb.sheetnames:
        ws = wb[name]
        for ci, h in enumerate(headers, 1): ws.cell(row=1, column=ci, value=h)
        style_header_row(ws, 1, len(headers))
        return ws
    ws = wb.create_sheet(name)
    for ci, h in enumerate(headers, 1): ws.cell(row=1, column=ci, value=h)
    style_header_row(ws, 1, len(headers))
    return ws

MAX_SUMMARY_ROWS = 365  # ~rok historii liczników w arkuszu profilu (1 wiersz/scan)

def _read_summary_rows(ws):
    """Czyta wiersze podsumowania (1/scan) z arkusza profilu.
    Wiersz podsumowania rozpoznajemy po liczbie w kolumnie 3 (Liczba ogłoszeń);
    wiersze ogłoszeń mają tu None, więc są pomijane."""
    rows = []
    for row in range(2, ws.max_row + 1):
        c3 = ws.cell(row=row, column=3).value
        if isinstance(c3, (int, float)):
            rows.append({
                "date":   ws.cell(row=row, column=1).value,
                "time":   ws.cell(row=row, column=2).value,
                "count":  int(c3),
                "change": ws.cell(row=row, column=4).value,
                "cross":  ws.cell(row=row, column=5).value,
            })
    return rows

def update_excel(scan_results, scan_timestamp):
    os.makedirs(DATA_DIR, exist_ok=True)
    wb = load_or_create_workbook()
    today   = scan_timestamp.strftime("%Y-%m-%d")
    now_str = scan_timestamp.strftime("%Y-%m-%d %H:%M")

    # Wczytaj stan JSON (refresh_count + price_history) — zapisany wcześniej przez
    # generate_dashboard_json (MUSI być wywołany przed update_excel).
    jd = {}
    refresh_count_map = {}
    if os.path.exists(JSON_PATH):
        try:
            with open(JSON_PATH, "r", encoding="utf-8") as f:
                jd = json.load(f)
            for pk, pd_ in jd.get("profiles", {}).items():
                for listing in pd_.get("current_listings", []):
                    refresh_count_map[listing.get("id","")] = listing.get("refresh_count", 0)
        except Exception:
            jd = {}

    profile_headers = [
        "Data scanu","Godzina","Liczba ogłoszeń","Zmiana vs poprzedni","Crosscheck",
        "Tytuł","Cena (zł)","Zmiana ceny","Promo","Dni Prom.","Sesje",
        "Data publikacji","Data odświeżenia","URL","Licz. odsw.","Licz. reakt.","Dni aktywne",
    ]

    for pk, result in scan_results.items():
        sheet_name = pk[:31]
        # Zachowaj dotychczasową serię liczników (1 wiersz/scan), potem przebuduj arkusz
        # od zera — dzięki temu snapshot ogłoszeń NIE kumuluje się między scanami.
        old_summary = _read_summary_rows(wb[sheet_name]) if sheet_name in wb.sheetnames else []
        prev_count  = old_summary[-1]["count"] if old_summary else None
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
        ws = get_or_create_sheet(wb, sheet_name, profile_headers)

        cur = result["count"]
        ch  = cur - prev_count if prev_count is not None else 0
        old_summary.append({"date": today, "time": scan_timestamp.strftime("%H:%M"),
                            "count": cur, "change": ch, "cross": result.get("crosscheck","")})
        if len(old_summary) > MAX_SUMMARY_ROWS:
            old_summary = old_summary[-MAX_SUMMARY_ROWS:]

        # Blok podsumowań — pełna seria liczników (ograniczona do MAX_SUMMARY_ROWS).
        srow = 1
        for s in old_summary:
            srow += 1
            ws.cell(row=srow, column=1, value=s["date"])
            ws.cell(row=srow, column=2, value=s["time"])
            ws.cell(row=srow, column=3, value=s["count"])
            sf = UP_FONT if (isinstance(s["change"],(int,float)) and s["change"] > 0) \
                 else DOWN_FONT if (isinstance(s["change"],(int,float)) and s["change"] < 0) else DATA_FONT
            style_data_cell(ws.cell(row=srow, column=4, value=s["change"]), sf)
            ws.cell(row=srow, column=5, value=s["cross"])
            for c in [1,2,3,5]: style_data_cell(ws.cell(row=srow, column=c))

        # Snapshot bieżących ogłoszeń — przebudowywany co scan (pusty wiersz separatora).
        snapshot_start = srow + 2
        for i, listing in enumerate(result["listings"]):
            row = snapshot_start + i
            pub, ref = parse_date_text(listing.get("date_text",""))
            lid = listing["listing_id"]
            is_promoted = listing.get("is_promoted", False)
            promo_days  = listing.get("promoted_days_current", 0)
            promo_sess  = listing.get("promoted_sessions_count", 0)
            refresh_cnt = refresh_count_map.get(lid, 0)
            react_cnt   = listing.get("reactivation_count", 0)

            # Days active
            first_seen_str = listing.get("first_seen","")
            try:
                fs = datetime.strptime(first_seen_str[:10], "%Y-%m-%d") if first_seen_str else None
                days_active = (datetime.now() - fs).days + 1 if fs else None
            except Exception: days_active = None

            ws.cell(row=row, column=1,  value=today)
            ws.cell(row=row, column=2,  value=scan_timestamp.strftime("%H:%M"))
            ws.cell(row=row, column=6,  value=listing["title"])
            ws.cell(row=row, column=7,  value=listing["price"])
            ws.cell(row=row, column=9,  value="★" if is_promoted else "")
            ws.cell(row=row, column=10, value=promo_days if is_promoted else None)
            ws.cell(row=row, column=11, value=promo_sess  if promo_sess > 0 else None)
            ws.cell(row=row, column=12, value=pub or "")
            ws.cell(row=row, column=13, value=ref or "")
            ws.cell(row=row, column=14, value=listing["url"])
            ws.cell(row=row, column=15, value=refresh_cnt)
            ws.cell(row=row, column=16, value=react_cnt)
            ws.cell(row=row, column=17, value=days_active)
            for c in range(1, 18): style_data_cell(ws.cell(row=row, column=c))

        widths = [12,8,15,15,14,50,12,12,8,10,8,14,14,60,10,10,10]
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    # Historia cen — odbudowywana z price_history (JSON = źródło prawdy zmian cen).
    # Rejestruje wyłącznie realne zmiany ceny; rozmiar ograniczony trymowaniem
    # price_history (90 dni) w generate_dashboard_json — arkusz już nie puchnie.
    ph = ["Data","Profil","ID ogłoszenia","Tytuł","Cena (zł)","Poprzednia cena","Zmiana ceny","URL"]
    if "historia_cen" in wb.sheetnames: del wb["historia_cen"]
    ws_p = get_or_create_sheet(wb, "historia_cen", ph)
    title_url = {}
    for pk, result in scan_results.items():
        for l in result["listings"]:
            title_url[l["listing_id"]] = (l["title"], l["url"])
    for pk_, pdata in jd.get("profiles", {}).items():
        for l in pdata.get("current_listings", []) + pdata.get("archived_listings", []):
            title_url.setdefault(l.get("id",""), (l.get("title",""), l.get("url","")))
    rp = 1
    for pk_, pdata in jd.get("profiles", {}).items():
        for lid, hist in pdata.get("price_history", {}).items():
            t, u = title_url.get(lid, ("",""))
            for h in hist:
                rp += 1
                chg = h.get("change")
                ws_p.cell(row=rp, column=1, value=h.get("date"))
                ws_p.cell(row=rp, column=2, value=pk_)
                ws_p.cell(row=rp, column=3, value=lid)
                ws_p.cell(row=rp, column=4, value=t)
                ws_p.cell(row=rp, column=5, value=h.get("new_price"))
                ws_p.cell(row=rp, column=6, value=h.get("old_price"))
                ws_p.cell(row=rp, column=7, value=chg)
                ws_p.cell(row=rp, column=8, value=u)
                for c in range(1,9):
                    cell = ws_p.cell(row=rp, column=c)
                    cf = DOWN_FONT if (c==7 and chg and chg<0) else UP_FONT if (c==7 and chg and chg>0) else DATA_FONT
                    style_data_cell(cell, cf)
    for idx, w in enumerate([18,18,15,50,12,14,12,60],1):
        ws_p.column_dimensions[get_column_letter(idx)].width = w

    # Podsumowanie
    if "podsumowanie" in wb.sheetnames: del wb["podsumowanie"]
    ws_s = wb.create_sheet("podsumowanie")
    sh = ["Profil","Label","Dzisiejsza liczba","Poprzednia liczba","Zmiana","Crosscheck","Data scanu"]
    for ci,h in enumerate(sh,1): ws_s.cell(row=1, column=ci, value=h)
    style_header_row(ws_s, 1, len(sh))
    ri = 2
    for pk, result in scan_results.items():
        cur = result["count"]; sn = pk[:31]; prev = None
        if sn in wb.sheetnames:
            counts = [int(wb[sn].cell(row=r,column=3).value) for r in range(2,wb[sn].max_row+1)
                      if wb[sn].cell(row=r,column=3).value is not None and isinstance(wb[sn].cell(row=r,column=3).value,(int,float))]
            if len(counts) >= 2: prev = counts[-2]
        ch = cur - prev if prev is not None else 0
        ws_s.cell(row=ri, column=1, value=pk)
        ws_s.cell(row=ri, column=2, value=PROFILES[pk]["label"])
        ws_s.cell(row=ri, column=3, value=cur)
        ws_s.cell(row=ri, column=4, value=prev)
        ws_s.cell(row=ri, column=5, value=ch)
        ws_s.cell(row=ri, column=6, value=result.get("crosscheck",""))
        ws_s.cell(row=ri, column=7, value=now_str)
        for c in range(1,8):
            cell = ws_s.cell(row=ri, column=c)
            f = UP_FONT if (c==5 and ch>0) else DOWN_FONT if (c==5 and ch<0) else DATA_FONT
            style_data_cell(cell, f)
        ri += 1
    for idx, w in enumerate([20,30,18,18,10,16,20],1):
        ws_s.column_dimensions[get_column_letter(idx)].width = w

    wb.save(EXCEL_PATH)
    log.info(f"Excel saved: {EXCEL_PATH}")

# ─── JSON for Dashboard ───────────────────────────────────────────────────────

def _days_since(ts_str, now_dt):
    """Ile dni minęło od znacznika "%Y-%m-%d %H:%M:%S" (None gdy nie da się sparsować)."""
    try:
        return (now_dt - datetime.strptime(str(ts_str)[:19], "%Y-%m-%d %H:%M:%S")).days
    except Exception:
        return None


def load_existing_json():
    if os.path.exists(JSON_PATH):
        try:
            with open(JSON_PATH,"r",encoding="utf-8") as f: return json.load(f)
        except (json.JSONDecodeError, IOError): pass
    return {"profiles":{}, "scan_history":[], "last_scan":None}

def build_price_distribution(listings):
    """Histogram cen aktywnych ofert (~14 słupków o „ładnym" kroku).
    Niezmiennik: suma count w słupkach == liczba ofert z dodatnią ceną."""
    prices = sorted([l["price"] for l in listings if l.get("price") and l["price"] > 0])
    if not prices:
        return []
    mn, mx = prices[0], prices[-1]
    if mn == mx:
        return [{"from": mn, "to": mx + 1, "count": len(prices)}]
    raw = (mx - mn) / 14
    mag = 10 ** int(len(str(int(raw))) - 1)
    step = next((f * mag for f in [1, 2, 2.5, 5, 10] if f * mag >= raw), 10 * mag)
    start = (mn // step) * step
    buckets = []
    s = start
    while s <= mx:
        cnt = sum(1 for p in prices if p >= s and p < s + step)
        buckets.append({"from": int(s), "to": int(s + step), "count": cnt})
        s += step
    # Przyciągnij puste krawędzie histogramu
    while len(buckets) > 1 and buckets[-1]["count"] == 0: buckets.pop()
    while len(buckets) > 1 and buckets[0]["count"] == 0:  buckets.pop(0)
    return buckets

def generate_dashboard_json(scan_results, scan_timestamp):
    data    = load_existing_json()
    now_str = scan_timestamp.strftime("%Y-%m-%d %H:%M:%S")
    today   = scan_timestamp.strftime("%Y-%m-%d")
    data["last_scan"] = now_str
    scan_entry = {"timestamp": now_str, "date": today, "profiles": {}}
    # Zbiera per-profil flow stats — zwracane do run_scan()
    profile_flow_stats = {}

    for pk, result in scan_results.items():
        cfg = PROFILES[pk]
        if pk not in data["profiles"]:
            data["profiles"][pk] = {
                "label": cfg["label"], "url": cfg["url"],
                "is_category": cfg.get("is_category", False),
                "daily_counts": [], "current_listings": [],
                "archived_listings": [], "price_history": {},
            }

        pd_ = data["profiles"][pk]
        dc  = pd_["daily_counts"]

        crosscheck   = result.get("crosscheck","")
        header_count = result.get("header_count")
        # Wynik bezpośredniego sprawdzenia ofert nieobecnych w sweepie (id → alive/dead/unknown).
        verification = result.get("verification") or {}
        # SAFETY: Każda z poniższych sytuacji oznacza uszkodzony scan, którego NIE wolno propagować:
        # 1) crosscheck == "error"  — wyjątek podczas scrapowania
        # 2) crosscheck == "anomaly_detected" — sanity check wykrył anomalię (count=0, spadek, ucięta paginacja...)
        # 3) count == 0 — pusty scan jest ZAWSZE błędem (niezależnie od header_count)
        is_scraper_error = (
            crosscheck in ("error", "anomaly_detected")
            or result["count"] == 0
        )
        current_listings_count = len(pd_.get("current_listings",[]))
        skip_daily_update = is_scraper_error and current_listings_count > 0

        if skip_daily_update:
            log.warning(f"[{pk}] Skipping update — scraper error (crosscheck={crosscheck}, count={result['count']}, header={header_count})")
            scan_entry["profiles"][pk] = {"count": result["count"], "crosscheck": crosscheck}
            if result.get("anomaly_reasons"):
                scan_entry["profiles"][pk]["anomaly_reasons"] = result["anomaly_reasons"]
            continue

        # ── Promoted & flow stats ──
        current_ids_new = {l["listing_id"] for l in result["listings"]}
        old_ids         = {l["id"] for l in pd_.get("current_listings",[])}
        newly_detected  = [l for l in result["listings"] if l["listing_id"] not in old_ids]

        # "Znikło" liczymy dopiero PO archiwizacji, jako len(newly_archived) — czyli
        # dokładnie te oferty, które w tym scanie zostały POTWIERDZONE jako usunięte:
        # bezpośrednim sprawdzeniem URL-a albo drugą nieobecnością z rzędu.
        first_run  = (len(old_ids) == 0 and len(dc) == 0)
        flow_added = None if first_run else len(current_ids_new - old_ids)

        total  = result["count"]
        promo_count = sum(1 for l in result["listings"] if l.get("is_promoted"))
        promo_pct   = round(promo_count / total * 100, 1) if total > 0 else 0

        # Price distribution snapshot (all active listings with price)
        price_dist = build_price_distribution(result["listings"])

        # Median from NEW listings only
        new_prices = [l["price"] for l in newly_detected if l.get("price") and l["price"] > 0]
        if new_prices:
            sp = sorted(new_prices); n = len(sp)
            median_price = sp[n//2] if n%2 != 0 else (sp[n//2-1]+sp[n//2])//2
        else:
            median_price = None

        # ── Build new_listings ──
        new_listings = []
        for listing in result["listings"]:
            pub, ref = parse_date_text(listing.get("date_text",""))
            nl = {
                "id": listing["listing_id"], "title": listing["title"],
                "price": listing["price"], "price_text": listing.get("price_text",""),
                "url": listing["url"], "published": pub, "refreshed": ref,
                "date_text": listing.get("date_text",""),
                "image_url": listing.get("image_url",""),
                "first_seen": now_str, "last_seen": now_str,
                "is_promoted": listing.get("is_promoted", False),
                "promotion_type": listing.get("promotion_type"),
                "refresh_count": 0,
                "promoted_days_current": 0,
                "promoted_sessions_count": 0,
                "promotion_history": [],
                "reactivation_count": 0,
                "reactivation_history": [],
            }
            new_listings.append(nl)

        old_map      = {l["id"]: l for l in pd_.get("current_listings",[])}
        archived_map = {l["id"]: l for l in pd_.get("archived_listings",[])}
        reactivated_ids = set()  # ID-ki przeniesione z archived → current; usuniemy je z archived na końcu

        for nl in new_listings:
            lid = nl["id"]
            if lid in old_map:
                old = old_map[lid]
                nl["first_seen"] = old.get("first_seen", now_str)

                # Price history
                old_price = old.get("price"); new_price = nl.get("price")
                if old_price is not None and new_price is not None and old_price != new_price:
                    if lid not in pd_["price_history"]: pd_["price_history"][lid] = []
                    pd_["price_history"][lid].append({
                        "date": now_str, "old_price": old_price,
                        "new_price": new_price, "change": new_price - old_price,
                    })
                    nl["previous_price"] = old_price
                    nl["price_change"]   = new_price - old_price
                elif lid in pd_.get("price_history",{}):
                    h = pd_["price_history"][lid]
                    if h:
                        nl["previous_price"] = h[-1]["old_price"]
                        nl["price_change"]   = (nl["price"] - h[-1]["old_price"]) if nl["price"] else None

                # Reactivation carry
                nl["reactivation_history"] = old.get("reactivation_history",[])
                nl["reactivation_count"]   = len(nl["reactivation_history"])

                # Refresh detection
                nl["refresh_count"]   = old.get("refresh_count", 0)
                nl["refresh_history"] = old.get("refresh_history",[])
                old_ref = old.get("refreshed")
                new_ref = nl.get("refreshed")
                if new_ref and new_ref != old_ref:
                    already = any(h.get("refreshed_at") == new_ref for h in nl["refresh_history"])
                    if not already:
                        nl["refresh_count"] += 1
                        nl["refresh_history"].append({
                            "refreshed_at": new_ref, "detected_at": now_str, "old_date": old_ref,
                        })
                        log.info(f"  [REFRESHED] {lid}: odświeżeń={nl['refresh_count']}")

                # Promotion tracking (historia per-ogłoszenie żyje w nl["promotion_history"])
                nl["promotion_history"]       = old.get("promotion_history",[])
                nl["promoted_days_current"]   = old.get("promoted_days_current", 0)
                nl["promoted_sessions_count"] = old.get("promoted_sessions_count", 0)
                old_promo = old.get("is_promoted", False)
                new_promo = nl.get("is_promoted", False)
                if new_promo and not old_promo:
                    nl["promotion_started_at"]    = now_str
                    nl["promoted_days_current"]   = 1
                    nl["promoted_sessions_count"] = old.get("promoted_sessions_count",0) + 1
                elif new_promo and old_promo:
                    nl["promotion_started_at"]  = old.get("promotion_started_at", now_str)
                    nl["promoted_days_current"] = old.get("promoted_days_current",0) + 1
                    nl["promoted_sessions_count"] = old.get("promoted_sessions_count",0)
                elif not new_promo and old_promo:
                    promo_start = old.get("promotion_started_at", now_str)
                    days = old.get("promoted_days_current",1)
                    nl["promotion_history"].append({
                        "start_date": promo_start, "end_date": now_str, "days": days,
                        "promotion_type": old.get("promotion_type","unknown"),
                        "session_number": old.get("promoted_sessions_count",0),
                    })
                    nl["promoted_days_current"] = 0
                    nl.pop("promotion_started_at", None)

            elif lid in archived_map:
                # Reactivation
                reactivated_ids.add(lid)
                old_archived = archived_map[lid]
                nl["first_seen"] = old_archived.get("first_seen", now_str)
                history = old_archived.get("reactivation_history",[])
                history.append({"active_from": old_archived.get("first_seen"), "reactivated_at": now_str})
                nl["reactivation_history"] = history
                nl["reactivation_count"]   = len(history)
                nl["refresh_count"]        = old_archived.get("refresh_count",0)
                nl["refresh_history"]      = old_archived.get("refresh_history",[])
                nl["promoted_days_current"]   = 0
                nl["promoted_sessions_count"] = old_archived.get("promoted_sessions_count",0)
                nl["promotion_history"]       = old_archived.get("promotion_history",[])
                if nl.get("is_promoted"):
                    nl["promotion_started_at"]    = now_str
                    nl["promoted_days_current"]   = 1
                    nl["promoted_sessions_count"] += 1
            else:
                # Brand new
                if nl.get("is_promoted"):
                    nl["promotion_started_at"]  = now_str
                    nl["promoted_days_current"] = 1
                    nl["promoted_sessions_count"] = 1

        # ── Archiwizacja ──
        # Kolejność werdyktów:
        #   1. weryfikacja po URL-u oferty — pomiar, nie domysł (alive / dead)
        #   2. gdy nierozstrzygnięta (403, timeout, nieznany layout) → stara heurystyka:
        #      archiwizacja dopiero przy drugiej nieobecności z rzędu.
        newly_archived  = []
        carried_missing = []   # nierozstrzygnięte, zostają w current_listings
        verified_alive  = []   # nieobecne w sweepie, ale OLX potwierdził, że żyją
        new_ids = {nl["id"] for nl in new_listings}

        def _archive(old_l, reason):
            old_l["archived_date"]   = now_str
            old_l["archived_reason"] = reason
            old_l.pop("missing_count", None)  # czyścimy przed archiwizacją
            r_hist = old_l.get("reactivation_history",[])
            if r_hist and "active_to_current" not in r_hist[-1]:
                r_hist[-1]["active_to_current"] = now_str
            old_l["reactivation_count"] = len(r_hist)
            if not old_l.get("refresh_history"): old_l["refresh_history"] = []
            if not old_l.get("refresh_count"):   old_l["refresh_count"] = len(old_l["refresh_history"])
            pd_["archived_listings"].append(old_l)
            newly_archived.append(old_l)

        for old_l in pd_.get("current_listings",[]):
            if old_l["id"] in new_ids:
                continue  # ogłoszenie zostało znalezione w tym scanie — obsłużone w new_listings
            status = verification.get(old_l["id"])

            if status == "alive":
                # Sweep ją zgubił, ale oferta żyje — zostaje aktywna, licznik nieobecności zerujemy.
                days_unseen = _days_since(old_l.get("last_seen"), scan_timestamp)
                if days_unseen is not None and days_unseen > VERIFY_MAX_ALIVE_DAYS:
                    # Bezpiecznik: gdyby detekcja martwej strony przestała działać,
                    # oferta trzymana wyłącznie weryfikacją nie może wisieć w nieskończoność.
                    log.warning(f"  [STALE-ALIVE] {old_l['id']}: żywa wg URL, ale nieobecna w sweepie "
                                f"od {days_unseen} dni — archiwizuję")
                    _archive(old_l, "stale_verified_alive")
                    continue
                old_l["missing_count"]     = 0
                old_l["last_verified"]     = now_str
                old_l["verified_alive_at"] = now_str
                verified_alive.append(old_l)
                log.info(f"  [VERIFIED-ALIVE] {old_l['id']}: {old_l.get('title','')[:50]}")
                continue

            if status == "dead":
                old_l["last_verified"] = now_str
                _archive(old_l, "verified_dead")
                log.info(f"  [ARCHIVED] {old_l['id']} (URL potwierdza usunięcie): {old_l.get('title','')[:50]}")
                continue

            prev_missing = int(old_l.get("missing_count", 0) or 0)
            new_missing = prev_missing + 1
            if new_missing >= 2:
                # Druga nieobecność z rzędu → archiwizacja
                _archive(old_l, "missing_2x")
                log.info(f"  [ARCHIVED] {old_l['id']} (missing 2× z rzędu): {old_l.get('title','')[:50]}")
            else:
                # Pierwsza nieobecność → trzymamy w current_listings z markerem
                old_l["missing_count"] = new_missing
                carried_missing.append(old_l)
                log.info(f"  [MISSING 1×] {old_l['id']}: {old_l.get('title','')[:50]}")

        # Archiwum zachowujemy w CAŁOŚCI i CIĄGLE — bez przycinania do ostatnich N.
        # Wcześniejszy cap (500) usuwał najstarsze wpisy, przez co historia odpływu
        # ofert (wykres „Odpływ ofert" w docs/trend.html) sięgała tylko ~14 dni wstecz.
        # Pełne archiwum = pełna historia zniknięć ofert z rynku.

        # ── Czyszczenie: usuń z archived te, które właśnie reaktywowano ──
        # Bez tego ta sama oferta istnieje w obu listach (current + archived) — bug do 2026-05-19.
        if reactivated_ids:
            before = len(pd_["archived_listings"])
            pd_["archived_listings"] = [
                l for l in pd_["archived_listings"] if l["id"] not in reactivated_ids
            ]
            removed = before - len(pd_["archived_listings"])
            if removed:
                log.info(f"  [DEDUPE] {pk}: usunięto {removed} reaktywowanych ofert z archived_listings")

        # Reset missing_count dla ogłoszeń, które wróciły (są w new_listings)
        for nl in new_listings:
            nl["missing_count"] = 0

        # ── daily_counts ──
        # `count`        — ile ofert zwrócił sweep (semantyka niezmieniona, żeby seria
        #                  historyczna dalej się zgadzała)
        # `active_count` — ile ofert uznajemy za żywe: znalezione w sweepie
        #                  + potwierdzone po URL-u + nierozstrzygnięte (czekają na drugą
        #                  nieobecność). To jest liczba "aktywnych ofert" dla dashboardu.
        flow_removed = None if first_run else len(newly_archived)
        active_count = len(new_listings) + len(verified_alive) + len(carried_missing)
        scan_meta = {
            "active_count":       active_count,
            "verified_alive":     len(verified_alive),
            "verified_dead":      sum(1 for l in newly_archived
                                      if l.get("archived_reason") == "verified_dead"),
            "unresolved_missing": len(carried_missing),
            "pages_scraped":      result.get("pages_scraped"),
            "pages_expected":     result.get("pages_expected"),
            "header_count":       header_count,
        }
        today_entry = next((d for d in dc if d["date"] == today), None)
        if today_entry:
            if result["count"] >= today_entry["count"]:
                today_entry["count"]               = result["count"]
                today_entry["timestamp"]           = now_str
                today_entry["median_price"]        = median_price
                today_entry["promoted_count"]      = promo_count
                today_entry["promoted_percentage"] = promo_pct
                today_entry["price_distribution"]  = price_dist
                today_entry.update(scan_meta)
                prev_added   = today_entry.get("added") or 0
                prev_removed = today_entry.get("removed") or 0
                if flow_added is not None:
                    today_entry["added"]   = prev_added + flow_added
                    today_entry["removed"] = prev_removed + flow_removed
                if len(dc) >= 2:
                    today_entry["change"] = result["count"] - dc[-2]["count"]
                    # Tylko wobec wpisu, który ma tę samą metrykę. Porównanie active_count
                    # z gołym count sprzed weryfikacji dawało skok będący artefaktem zmiany
                    # definicji (scan #144: +99 zamiast realnych +36).
                    prev_active = dc[-2].get("active_count")
                    today_entry["active_change"] = (active_count - prev_active
                                                    if prev_active is not None else None)
        else:
            prev_c      = dc[-1]["count"] if dc else None
            prev_active = dc[-1].get("active_count") if dc else None
            ch          = result["count"] - prev_c if prev_c is not None else 0
            entry = {
                "date": today, "count": result["count"], "change": ch,
                # None = brak porównywalnej podstawy (poprzedni wpis sprzed weryfikacji);
                # front spada wtedy na `change`, zamiast pokazywać artefakt zmiany metryki.
                "active_change": (active_count - prev_active) if prev_active is not None else None,
                "timestamp": now_str, "median_price": median_price,
                "promoted_count": promo_count, "promoted_percentage": promo_pct,
                "price_distribution": price_dist,
                "refreshed_count": 0, "reactivated_count": 0,
                "added": flow_added, "removed": flow_removed,
            }
            entry.update(scan_meta)
            dc.append(entry)
        if len(dc) > 90: pd_["daily_counts"] = dc[-90:]

        # ── Count refreshes & reactivations today ──
        reactivated_count = 0; refreshed_count = 0
        for l in list(new_listings) + newly_archived:
            rh = l.get("reactivation_history",[])
            if rh and rh[-1].get("reactivated_at","").startswith(today): reactivated_count += 1
            fh = l.get("refresh_history",[])
            if fh and fh[-1].get("detected_at","").startswith(today): refreshed_count += 1
        te = next((d for d in dc if d["date"] == today), None)
        if te:
            te["reactivated_count"] = reactivated_count
            te["refreshed_count"]   = refreshed_count

        # current_listings = znalezione w sweepie + potwierdzone po URL-u + nierozstrzygnięte
        pd_["current_listings"] = new_listings + verified_alive + carried_missing

        # ── Zapora: masowy "missing" w jednym scanie ──────────────────────
        # Gdy duża część bazy znika w pojedynczym scanie (carried_missing), to
        # niemal zawsze niepełny scrape OLX, a nie realny odpływ ofert. Spadek
        # total bywa poniżej progu SANITY_MAX_DROP_RATIO (bo "missing" != "removed",
        # archiwizacja jest 2-scan), więc zwykła zapora go NIE łapie. Sygnalizujemy
        # to przez API — operator dostaje ostrzeżenie już po pierwszym niepełnym
        # scanie, zanim te oferty zostaną (przy drugim z rzędu) masowo zarchiwizowane.
        base_for_missing = current_listings_count   # liczba ofert sprzed tego scanu
        freshly_missing  = len(carried_missing)
        missing_ratio    = (freshly_missing / base_for_missing) if base_for_missing else 0.0
        partial_scan_warning = None
        if base_for_missing and missing_ratio > SANITY_MAX_MISSING_RATIO:
            partial_scan_warning = {
                "missing_this_scan": freshly_missing,
                "base_count":        base_for_missing,
                "scanned_count":     result["count"],
                "missing_ratio":     round(missing_ratio, 3),
                "verification_skipped": (result.get("verification_stats") or {}).get("skipped"),
                "message": (f"{freshly_missing} z {base_for_missing} ofert "
                            f"({missing_ratio*100:.1f}%) zniknęło w jednym scanie — "
                            f"prawdopodobnie niepełny scrape OLX, nie realny odpływ"),
            }
            log.warning(f"[PARTIAL-SCAN] {pk}: {partial_scan_warning['message']}")

        # Trymowanie price_history do ostatnich 90 dni — ogranicza wzrost JSON.
        # Porównanie stringów "%Y-%m-%d %H:%M:%S" jest chronologiczne leksykograficznie.
        cutoff = (scan_timestamp - timedelta(days=90)).strftime("%Y-%m-%d %H:%M:%S")
        ph_map = pd_.get("price_history", {})
        for ph_lid in list(ph_map.keys()):
            kept = [h for h in ph_map[ph_lid] if h.get("date", "") >= cutoff]
            if kept:
                ph_map[ph_lid] = kept
            else:
                del ph_map[ph_lid]

        scan_entry["profiles"][pk] = {
            "count": result["count"], "active_count": active_count, "crosscheck": crosscheck,
            "verified_alive": scan_meta["verified_alive"], "verified_dead": scan_meta["verified_dead"],
        }
        if partial_scan_warning:
            scan_entry["profiles"][pk]["partial_scan_warning"] = partial_scan_warning

        # Zapisz flow stats per profil
        vstats = result.get("verification_stats") or {}
        profile_flow_stats[pk] = {
            "label":            PROFILES[pk]["label"],
            "listings_total":   result["count"],       # znalezione w sweepie (bez zmian semantyki)
            "listings_active":  active_count,          # uznane za żywe (nowa metryka)
            "listings_new":     flow_added,
            "listings_removed": flow_removed,
            "crosscheck":       crosscheck,
            "verified_alive":   scan_meta["verified_alive"],
            "verified_dead":    scan_meta["verified_dead"],
            "verified_unknown": vstats.get("unknown"),
            "verification_skipped": vstats.get("skipped"),
            "pages_scraped":    result.get("pages_scraped"),
            "pages_expected":   result.get("pages_expected"),
            "header_count":     header_count,
            "partial_scan_warning": partial_scan_warning,
        }

    data["scan_history"].append(scan_entry)
    if len(data["scan_history"]) > 90: data["scan_history"] = data["scan_history"][-90:]
    with open(JSON_PATH,"w",encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    log.info(f"Dashboard JSON saved: {JSON_PATH}")
    return profile_flow_stats

# ─── Main ─────────────────────────────────────────────────────────────────────

def run_scan():
    ts = datetime.now()
    log.info(f"{'='*60}")
    log.info(f"SZPERACZ MIESZKANIOWY — Scan started {ts.strftime('%Y-%m-%d %H:%M:%S')}")
    log.info(f"{'='*60}")
    results = {}
    for pk, cfg in PROFILES.items():
        try:
            r = scrape_with_crosscheck(pk, cfg)
            # Oferty z bazy nieobecne w sweepie sprawdzamy bezpośrednio po ich URL-u —
            # dopiero z tym werdyktem generate_dashboard_json decyduje o archiwizacji.
            verify_missing_for_profile(pk, r)
            results[pk] = r
            log.info(f"[OK] {pk}: {r['count']} listings ({r['crosscheck']})")
        except Exception as e:
            log.error(f"[ERROR] {pk}: {e}")
            results[pk] = {"listings":[], "count":0, "header_count":None, "crosscheck":"error", "pages_scraped":0}
        time.sleep(random.uniform(2,4))
    profile_flow = generate_dashboard_json(results, ts)
    # Doklejamy flow stats do results — używane przez main.py do scan_status.json
    for pk, flow in profile_flow.items():
        if pk in results:
            results[pk]["flow"] = flow
    update_excel(results, ts)
    log.info(f"SZPERACZ MIESZKANIOWY — Scan completed {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    return results

if __name__ == "__main__":
    run_scan()
